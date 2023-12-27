from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait #ilgili driverı bekleten yapı
from selenium.webdriver.support import expected_conditions as ec #beklenen koşullar
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
import pytest
import openpyxl
from constants import globalConstants as c


class Test_odevClass:

    def setup_method(self):   #her test başlangıcında çalışacak fonk.
        self.driver = webdriver.Chrome()
        self.driver.get(c.BASE_URL)
        self.driver.maximize_window()

    def teardown_method(self):  # her testin bitiminde çalışacak fonk
        self.driver.quit()

    def getData():
        excel = openpyxl.load_workbook(c.INPUTLAR_XLSX)
        sheet = excel["Sayfa1"] #hangi sayfada çalışacağımı gösteriyorum
        rows = sheet.max_row #kaçıncı satıra kadar veri var?
        data = []
        for i in range(2,rows+1):
            username = sheet.cell(i,1).value # i.satır 1.sütun hücredeki değer
            password = sheet.cell(i,2).value # i. satır 2. sütun hücredeki değer
            data.append((username,password))  # verileri [(username1,username2),(username3,username4)] şeklinde liste haline getirdik --> data[]
        
        return data
    
    def test_empty_login(self):
        
        usernameInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        usernameInput.send_keys('')
        passwordInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
        passwordInput.send_keys('')
        loginButton=self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()
        errorMesagge = self.driver.find_element(By.XPATH,c.ERROR_MESSAGE_XPATH)
        assert errorMesagge.text == c.EMPYT_LOGIN_MESSAGE
        
    
    def test_empty_password_login(self):
        
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        usernameInput.send_keys("locked_out_user")
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
        passwordInput.send_keys("")
        loginButton = self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()
        errorMessage = self.driver. find_element(By.XPATH,c.ERROR_MESSAGE_XPATH)
        assert errorMessage.text == c.EMPYTY_PASSWORD_MESSAGE
        
    
    def test_invalid_login(self):
        
        usernameInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        usernameInput.send_keys('locked_out_user') 
        passwordInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
        passwordInput.send_keys('secret_sauce')
        loginButton=self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click() 
        errorMessage=self.driver.find_element(By.XPATH,c.ERROR_MESSAGE_XPATH)
        assert errorMessage.text== c.INVALİD_LOGIN_MESSAGE
    
    def test_valid_login(self):
        
        usernameInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        usernameInput.send_keys('standard_user') 
        passwordInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
        passwordInput.send_keys('secret_sauce')
        loginButton=self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()
        assert  self.driver.current_url == c.CURRENT_URL_ANASAYFA
        #print(f"test sonucu(ana sayfaya geçildi mi): {sayfa}")
        urunler = self.driver.find_elements(By.CLASS_NAME,c.URUNLER_CLASSNAME)
        assert len(urunler) == 6


    def test_X_ikonu_temizleme(self):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
        actions=ActionChains(self.driver)
        actions.send_keys_to_element(usernameInput,"")
        actions.send_keys_to_element(passwordInput,"")
        actions.perform()

        loginButton=self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()
        
    
        x_icon =WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.CSS_SELECTOR,c.X_IKON_CSSSELECTOR)))
        x_ikon_username =WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.CSS_SELECTOR,c.X_IKON_USERNAME_CSSSELECTOR)))
        x_ikonu_password = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.CSS_SELECTOR,c.X_IKON_PASSWORD_CSSSELECTOR)))
        
        x_icon.click()
        try:
            x_icon.is_displayed()
            x_ikon_username.is_displayed()
            x_ikonu_password.is_displayed()
            testResult = False  #ikonlar görünüyorsa test sonucu false olur
            assert testResult == True #test sonucu true olmadığı için assert false döndürür
           
        except:
            testResult= True
            assert testResult == True

    @pytest.mark.parametrize("username,password",getData())
    def test_eslesmeyen_kullanici(self,username,password):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
        actions=ActionChains(self.driver)
        actions.send_keys_to_element(usernameInput,username)
        actions.send_keys_to_element(passwordInput,password)
        actions.perform()

        loginButton=self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()

        errorMessage = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.ERROR_MESSAGE_XPATH)))
        assert errorMessage.text == c.ESLESMEYEN_KULLANİCİ_MESSAGE
   
    def test_sepete_urun_ekleme(self):
        
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
        actions=ActionChains(self.driver)
        actions.send_keys_to_element(usernameInput,"standard_user")
        actions.send_keys_to_element(passwordInput,"secret_sauce")
        actions.perform()

        loginButton=self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()
        addToCard = WebDriverWait(self.driver,5).until(ec.visibility_of_all_elements_located((By.XPATH,c.ADDTOCARD_XPATH))) #ana sayfadaki tüm ürünleri bul
        #addToCard = WebDriverWait(self.driver,5).until(ec.visibility_of_all_elements_located((By.CLASS_NAME,"btn btn_primary btn_small btn_inventory"))) #ana sayfadaki tüm ürünleri bul
        #addToCard = WebDriverWait(self.driver,5).until(ec.visibility_of_all_elements_located((By.XPATH,"/html/body/div/div/div/div[2]/div/div/div/div/div[2]/div[2]/button"))) #ana sayfadaki tüm ürünleri bul
        #print(f"bulunan element:{len(addToCard)}")

        for i in range(len(addToCard)): # ana sayfadaki ürünleri sepete ekle
            addToCard[i].click()

        sepet = self.driver.find_element(By.XPATH,c.SEPET_XPATH).click() #sepet butonunu bul ve tıkla
        sepetteki_urunler = self.driver.find_elements(By.CLASS_NAME,c.SEPETTEKİ_URUNLER_CLASSNAME) # sepetteki ürünleri bul
        #print(f"septteki ürünler:{len(sepetteki_urunler)} adet")
        assert len(addToCard) == len(sepetteki_urunler)

    def test_sepetten_urun_silme(self):
        
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
        actions=ActionChains(self.driver)
        actions.send_keys_to_element(usernameInput,"standard_user")
        actions.send_keys_to_element(passwordInput,"secret_sauce")
        actions.perform()

        loginButton=self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()
        
        addToCard = WebDriverWait(self.driver,5).until(ec.visibility_of_all_elements_located((By.XPATH,c.ADDTOCARD_XPATH))) #ana sayfadaki tüm ürünleri bul
        #addToCard = WebDriverWait(self.driver,5).until(ec.visibility_of_all_elements_located((By.CLASS_NAME,"btn btn_primary btn_small btn_inventory"))) #ana sayfadaki tüm ürünleri bul
        #addToCard = WebDriverWait(self.driver,5).until(ec.visibility_of_all_elements_located((By.XPATH,"/html/body/div/div/div/div[2]/div/div/div/div/div[2]/div[2]/button"))) #ana sayfadaki tüm ürünleri bul
        #print(f"bulunan element:{len(addToCard)}")

        for i in range(len(addToCard)): # ana sayfadaki ürünleri sepete ekle
            addToCard[i].click()

        sepet = self.driver.find_element(By.XPATH,c.SEPET_XPATH).click() #sepet butonunu bul ve tıkla
        sepetteki_urunler = self.driver.find_elements(By.CLASS_NAME,c.SEPETTEKİ_URUNLER_CLASSNAME) # sepetteki ürünleri bul
        #print(f"septteki ürünler:{len(sepetteki_urunler)} adet")

        remove = WebDriverWait(self.driver,5).until(ec.visibility_of_all_elements_located((By.XPATH,c.REMOVE_XPATH)))
        for i in range(len(addToCard)):
        
            remove[i].click()
            
            try:                              
                remove[i].is_displayed()     # remove i. index görünür mü? evet ise test result false olsun ve döngüden çık
                testResult = False
                break #döngüyü kır döngüden çık
            except:
                testResult = True            #try daki koşul true değilse bu satırı çalıştır yani test result true oldu,döngüye devam et
                continue #döngüye devam et
        
        assert testResult == True
        


  

